"""
AttendIQ — Export Manager.
Generates CSV and Excel attendance reports for a lecturer's classes.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import os
from datetime import datetime, timedelta
import logging

from gui.components import (
    PALETTE, FONTS, Card, DataTable, Toast,
)

logger = logging.getLogger(__name__)


class ExportManagerPanel(ttk.Frame):

    def __init__(self, parent, db, user: dict, root_ref):
        super().__init__(parent, style="TFrame")
        self._db      = db
        self._user    = user
        self._root    = root_ref
        self._classes: dict = {}
        self._preview_data: list[dict] = []
        self._build()
        self._refresh_classes()

    def _build(self):
        # Options panel
        opt_card = Card(self, "Export Options")
        opt_card.pack(fill=tk.X, pady=(0, 12), padx=4)

        grid = tk.Frame(opt_card, bg=PALETTE["SURFACE"])
        grid.pack(fill=tk.X)

        # Row 0: Class
        tk.Label(grid, text="Class", font=FONTS["LABEL"],
                 bg=PALETTE["SURFACE"], fg=PALETTE["SUBTEXT"]).grid(
            row=0, column=0, sticky="w", pady=4, padx=(0, 10))
        self._class_var = tk.StringVar()
        self._class_cb  = ttk.Combobox(grid, textvariable=self._class_var,
                                        state="readonly", width=36)
        self._class_cb.grid(row=0, column=1, sticky="ew", pady=4)

        # Row 1: From date
        tk.Label(grid, text="From", font=FONTS["LABEL"],
                 bg=PALETTE["SURFACE"], fg=PALETTE["SUBTEXT"]).grid(
            row=1, column=0, sticky="w", pady=4, padx=(0, 10))
        try:
            from tkcalendar import DateEntry
            default_from = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            self._from_entry = DateEntry(grid, width=14, date_pattern="yyyy-mm-dd",
                                          background=PALETTE["ACCENT"],
                                          foreground=PALETTE["WHITE"])
        except ImportError:
            default_from = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
            self._from_var   = tk.StringVar(value=default_from)
            self._from_entry = ttk.Entry(grid, textvariable=self._from_var, width=14)
        self._from_entry.grid(row=1, column=1, sticky="w", pady=4)

        # Row 2: To date
        tk.Label(grid, text="To", font=FONTS["LABEL"],
                 bg=PALETTE["SURFACE"], fg=PALETTE["SUBTEXT"]).grid(
            row=2, column=0, sticky="w", pady=4, padx=(0, 10))
        try:
            from tkcalendar import DateEntry
            self._to_entry = DateEntry(grid, width=14, date_pattern="yyyy-mm-dd",
                                        background=PALETTE["ACCENT"],
                                        foreground=PALETTE["WHITE"])
        except ImportError:
            self._to_var   = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
            self._to_entry = ttk.Entry(grid, textvariable=self._to_var, width=14)
        self._to_entry.grid(row=2, column=1, sticky="w", pady=4)

        # Row 3: Format
        tk.Label(grid, text="Format", font=FONTS["LABEL"],
                 bg=PALETTE["SURFACE"], fg=PALETTE["SUBTEXT"]).grid(
            row=3, column=0, sticky="w", pady=4, padx=(0, 10))
        self._fmt_var = tk.StringVar(value="csv")
        fmt_row = tk.Frame(grid, bg=PALETTE["SURFACE"])
        fmt_row.grid(row=3, column=1, sticky="w", pady=4)
        ttk.Radiobutton(fmt_row, text="CSV",
                        variable=self._fmt_var, value="csv").pack(side=tk.LEFT)
        ttk.Radiobutton(fmt_row, text="Excel (.xlsx)",
                        variable=self._fmt_var, value="xlsx").pack(side=tk.LEFT, padx=(16, 0))

        grid.columnconfigure(1, weight=1)

        # Action buttons
        btn_row = tk.Frame(opt_card, bg=PALETTE["SURFACE"])
        btn_row.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(btn_row, text="Preview",
                   command=self._preview).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="⬇ Export Report",
                   style="Primary.TButton",
                   command=self._export).pack(side=tk.LEFT)
        ttk.Button(btn_row, text="📊 Full Summary Export",
                   command=self._export_summary).pack(side=tk.LEFT, padx=(8, 0))

        # Preview table
        preview_card = Card(self, "Preview")
        preview_card.pack(fill=tk.BOTH, expand=True, padx=4)

        prev_cols = [
            {"key": "sid",      "label": "Student ID",      "width": 110},
            {"key": "name",     "label": "Name",            "width": 180, "stretch": True},
            {"key": "present",  "label": "Sessions Present","width": 130, "anchor": "center"},
            {"key": "total",    "label": "Total Sessions",  "width": 120, "anchor": "center"},
            {"key": "percent",  "label": "Attendance %",    "width": 110, "anchor": "center"},
        ]
        self._preview_table = DataTable(preview_card, columns=prev_cols, height=16)
        self._preview_table.pack(fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _refresh_classes(self):
        classes = self._db.get_classes(lecturer_id=self._user["id"])
        self._classes = {f"{c['code']} — {c['name']}": c for c in classes}
        self._class_cb["values"] = list(self._classes.keys())
        if self._classes:
            self._class_cb.current(0)

    def _get_dates(self) -> tuple[str, str]:
        try:
            date_from = self._from_entry.get_date().strftime("%Y-%m-%d")
        except AttributeError:
            date_from = self._from_var.get()
        try:
            date_to = self._to_entry.get_date().strftime("%Y-%m-%d")
        except AttributeError:
            date_to = self._to_var.get()
        return date_from, date_to

    def _get_class(self) -> dict | None:
        key = self._class_var.get()
        return self._classes.get(key)

    def _preview(self):
        cls = self._get_class()
        if not cls:
            Toast(self._root, "Select a class.", "warning")
            return
        date_from, date_to = self._get_dates()
        self._preview_data = self._db.get_full_report(cls["id"], date_from, date_to)

        rows, tags = [], []
        for d in self._preview_data:
            pct = d["percent"]
            tag = "success" if pct >= 75 else ("warning" if pct >= 60 else "danger")
            rows.append((
                d.get("student_id") or "—",
                d["full_name"],
                d["sessions_present"],
                d["total_sessions"],
                f"{pct:.1f}%",
            ))
            tags.append(tag)

        self._preview_table.load(rows, tags)

    def _export(self):
        cls = self._get_class()
        if not cls:
            Toast(self._root, "Select a class.", "warning")
            return

        self._preview()  # ensure data is fresh
        if not self._preview_data:
            Toast(self._root, "No data to export.", "info")
            return

        date_from, date_to = self._get_dates()
        fmt = self._fmt_var.get()
        default_name = (f"{cls['code']}_attendance_"
                        f"{date_from}_to_{date_to}.{fmt}")

        if fmt == "csv":
            self._export_csv(cls, date_from, date_to, default_name)
        else:
            self._export_xlsx(cls, date_from, date_to, default_name)

    def _export_csv(self, cls: dict, date_from: str, date_to: str, default_name: str):
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=default_name,
        )
        if not path:
            return

        records = self._db.get_attendance_range(cls["id"], date_from, date_to)

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                # Header block
                w.writerow([f"Class: {cls['code']} — {cls['name']}"])
                w.writerow([f"Lecturer: {self._user['full_name']}"])
                w.writerow([f"Period: {date_from} to {date_to}"])
                w.writerow([f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
                w.writerow([])
                w.writerow(["--- Per-Session Records ---"])
                w.writerow(["Date", "Student ID", "Name", "Time", "Method", "Confidence"])
                for r in records:
                    conf = f"{r['confidence']:.4f}" if r.get("confidence") else "manual"
                    w.writerow([r["session_date"],
                                r.get("student_number") or "—",
                                r["full_name"],
                                r["timestamp"],
                                r["method"],
                                conf])
                w.writerow([])
                w.writerow(["--- Summary ---"])
                w.writerow(["Student ID", "Name", "Sessions Present", "Total Sessions", "Attendance %"])
                for d in self._preview_data:
                    w.writerow([d.get("student_id") or "—",
                                d["full_name"],
                                d["sessions_present"],
                                d["total_sessions"],
                                f"{d['percent']:.1f}%"])

            Toast(self._root, f"Exported: {os.path.basename(path)}", "success")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_xlsx(self, cls: dict, date_from: str, date_to: str, default_name: str):
        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "Install openpyxl to export Excel files:\n  pip install openpyxl"
            )
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_name,
        )
        if not path:
            return

        wb = openpyxl.Workbook()

        # -- Sheet 1: Summary
        ws = wb.active
        ws.title = "Summary"
        hdr_fill   = PatternFill("solid", fgColor="1a1a2e")
        hdr_font   = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, color="e94560", size=13)
        thin = Border(bottom=Side(style="thin", color="2a2a4a"))

        ws["A1"] = f"AttendIQ — {cls['code']} Attendance Report"
        ws["A1"].font = title_font
        ws["A2"] = f"Lecturer: {self._user['full_name']}"
        ws["A3"] = f"Period: {date_from}  →  {date_to}"
        ws["A4"] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws.append([])

        headers = ["Student ID", "Name", "Sessions Present",
                   "Total Sessions", "Attendance %"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for d in self._preview_data:
            pct = d["percent"]
            row = [d.get("student_id") or "—",
                   d["full_name"],
                   d["sessions_present"],
                   d["total_sessions"],
                   f"{pct:.1f}%"]
            ws.append(row)
            # Colour % cell
            pct_cell = ws.cell(ws.max_row, 5)
            if pct >= 75:
                pct_cell.fill = PatternFill("solid", fgColor="0d2a1e")
                pct_cell.font = Font(color="16c79a", bold=True)
            elif pct >= 60:
                pct_cell.fill = PatternFill("solid", fgColor="2a1e00")
                pct_cell.font = Font(color="f5a623", bold=True)
            else:
                pct_cell.fill = PatternFill("solid", fgColor="2a0d15")
                pct_cell.font = Font(color="e94560", bold=True)

        for col in ws.columns:
            width = max(len(str(c.value or "")) + 4 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(width, 40)

        # -- Sheet 2: Raw records
        ws2 = wb.create_sheet("Raw Records")
        ws2.append(["Date", "Student ID", "Name", "Time", "Method", "Confidence"])
        for cell in ws2[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        records = self._db.get_attendance_range(cls["id"], date_from, date_to)
        for r in records:
            conf = round(r["confidence"], 4) if r.get("confidence") else "manual"
            ws2.append([r["session_date"],
                        r.get("student_number") or "—",
                        r["full_name"],
                        r["timestamp"],
                        r["method"],
                        conf])

        try:
            wb.save(path)
            Toast(self._root, f"Excel exported: {os.path.basename(path)}", "success")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _export_summary(self):
        """Export one sheet per class across all of the lecturer's classes."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("Missing Dependency",
                                  "Install openpyxl:  pip install openpyxl")
            return

        date_from, date_to = self._get_dates()
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"full_report_{date_from}_to_{date_to}.xlsx",
        )
        if not path:
            return

        wb   = openpyxl.Workbook()
        first = True
        classes = self._db.get_classes(lecturer_id=self._user["id"])

        for cls in classes:
            ws = wb.active if first else wb.create_sheet()
            ws.title = cls["code"][:31]
            first = False

            ws["A1"] = f"{cls['code']} — {cls['name']}"
            ws["A1"].font = Font(bold=True, size=12, color="e94560")
            ws["A2"] = f"Period: {date_from} to {date_to}"
            ws.append([])

            ws.append(["Student ID", "Name", "Present", "Total", "%"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1a1a2e")

            summary = self._db.get_full_report(cls["id"], date_from, date_to)
            for d in summary:
                ws.append([d.get("student_id") or "—",
                            d["full_name"],
                            d["sessions_present"],
                            d["total_sessions"],
                            f"{d['percent']:.1f}%"])

        try:
            wb.save(path)
            Toast(self._root, "Full summary exported.", "success")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))
